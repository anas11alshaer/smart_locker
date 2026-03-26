"""Bulk device import from Excel (.xlsx) files.

Reads devices from an Excel sheet and inserts them into the database,
skipping duplicates (by PM number).

Usage:
    python -m scripts.import_devices --file devices.xlsx

    # With custom column mapping (if your headers differ):
    python -m scripts.import_devices --file devices.xlsx --pm-col "Equipment" --serial-col "S/N"

    # With optional type and slot columns:
    python -m scripts.import_devices --file devices.xlsx --type-col "Kategorie" --slot-col "Platz"

    # Specify which sheet to read (default: first sheet):
    python -m scripts.import_devices --file devices.xlsx --sheet "Sheet2"

    # Dry run — preview what would be imported without writing to DB:
    python -m scripts.import_devices --file devices.xlsx --dry-run
"""

import argparse
import sys
from datetime import date, datetime
from pathlib import Path

# Add project root to path
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from openpyxl import load_workbook

from config.logging_config import setup_logging
from smart_locker.database.engine import get_session, init_db
from smart_locker.database.repositories import DeviceRepository


def _find_column(headers: list[str], candidates: list[str]) -> int | None:
    """Find a column index by trying multiple possible header names (case-insensitive)."""
    lower_candidates = [c.lower() for c in candidates]
    for i, header in enumerate(headers):
        if header and header.strip().lower() in lower_candidates:
            return i
    return None


def _parse_date(value) -> date | None:
    """Parse a date from an Excel cell value (datetime object or string)."""
    if value is None:
        return None
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()
    # Try common date string formats (German DD.MM.YYYY first)
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Bulk import devices from an Excel file."
    )
    parser.add_argument("--file", required=True, help="Path to the .xlsx file")
    parser.add_argument("--sheet", default=None, help="Sheet name (default: first sheet)")
    # Column mapping overrides
    parser.add_argument("--pm-col", default=None, help="Column header for PM/equipment number")
    parser.add_argument("--name-col", default=None, help="Column header for device name (overrides auto-composition)")
    parser.add_argument("--serial-col", default=None, help="Column header for serial number")
    parser.add_argument("--type-col", default=None, help="Column header for device type/category")
    parser.add_argument("--slot-col", default=None, help="Column header for locker slot")
    parser.add_argument("--manufacturer-col", default=None, help="Column header for manufacturer")
    parser.add_argument("--model-col", default=None, help="Column header for model/type designation")
    parser.add_argument("--barcode-col", default=None, help="Column header for barcode")
    parser.add_argument("--calibration-col", default=None, help="Column header for calibration due date")
    # Defaults
    parser.add_argument(
        "--default-type", default="general", help="Default device type if no type column (default: general)"
    )
    parser.add_argument("--dry-run", action="store_true", help="Preview without writing to database")
    args = parser.parse_args()

    setup_logging()

    # Load Excel file
    file_path = Path(args.file)
    if not file_path.exists():
        print(f"ERROR: File not found: {file_path}")
        return

    print(f"Reading: {file_path}")
    wb = load_workbook(file_path, read_only=True, data_only=True)

    if args.sheet:
        if args.sheet not in wb.sheetnames:
            print(f"ERROR: Sheet '{args.sheet}' not found. Available: {wb.sheetnames}")
            return
        ws = wb[args.sheet]
    else:
        ws = wb.active
    print(f"Sheet: {ws.title}")

    # Read all rows
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 2:
        print("ERROR: Sheet has no data rows (need header + at least one data row).")
        return

    # Parse headers (first row)
    headers = [str(h).strip() if h else "" for h in rows[0]]
    print(f"Headers found: {headers}")

    # --- Column auto-detection candidates ---
    pm_candidates = [
        "equipment", "pm number", "pm", "pm_number", "equipment number",
        "equipmentnumber", "inventarnummer",
    ]
    name_candidates = [
        "name", "device name", "device", "equipment name", "item", "item name",
    ]
    serial_candidates = [
        "serial", "serial number", "s/n", "sn", "serial no", "serial_number",
        "serialnumber", "hersteller-serialnummer", "herstellerseriennummer",
        "seriennummer",
    ]
    type_candidates = [
        "type", "device type", "category", "device_type", "kind",
        "kategorie",
    ]
    slot_candidates = [
        "slot", "locker slot", "locker", "locker_slot", "bay",
        "platz messmittelschrank", "platz", "messmittelschrank", "schrank",
    ]
    desc_candidates = [
        "description", "desc", "details", "notes",
        "beschreibung",
    ]
    image_candidates = [
        "image", "photo", "image_path", "photo_path", "img", "picture", "filename",
    ]
    manufacturer_candidates = [
        "manufacturer", "make", "brand", "hersteller",
    ]
    model_candidates = [
        "model", "model name", "type designation",
        "typbezeichnung", "typ", "modell",
    ]
    barcode_candidates = [
        "barcode", "bar code", "barcodenummer",
    ]
    calibration_candidates = [
        "calibration due", "calibration_due", "next calibration",
        "datum der nächsten kalibrierung", "datum der nachsten kalibrierung",
        u"datum der n\u00e4chsten kalibrierung",
        "nächste kalibrierung", "kalibrierung", "kalibrierdatum",
    ]

    # --- Detect columns ---
    pm_idx = _find_column(headers, [args.pm_col] if args.pm_col else pm_candidates)
    name_idx = _find_column(headers, [args.name_col] if args.name_col else name_candidates)
    serial_idx = _find_column(headers, [args.serial_col] if args.serial_col else serial_candidates)
    type_idx = _find_column(headers, [args.type_col] if args.type_col else type_candidates)
    slot_idx = _find_column(headers, [args.slot_col] if args.slot_col else slot_candidates)
    desc_idx = _find_column(headers, desc_candidates)
    image_idx = _find_column(headers, image_candidates)
    mfr_idx = _find_column(headers, [args.manufacturer_col] if args.manufacturer_col else manufacturer_candidates)
    model_idx = _find_column(headers, [args.model_col] if args.model_col else model_candidates)
    barcode_idx = _find_column(headers, [args.barcode_col] if args.barcode_col else barcode_candidates)
    cal_idx = _find_column(headers, [args.calibration_col] if args.calibration_col else calibration_candidates)

    # PM number is required
    if pm_idx is None:
        print(f"ERROR: Could not find a PM/equipment number column. Headers: {headers}")
        print("Use --pm-col to specify the column header for PM number.")
        return

    # Name will be auto-composed from PM + manufacturer + model if no name column found
    compose_name = name_idx is None

    # Print column mapping
    print(f"\nColumn mapping:")
    print(f"  PM number: '{headers[pm_idx]}' (col {pm_idx + 1})")
    if name_idx is not None:
        print(f"  Name: '{headers[name_idx]}' (col {name_idx + 1})")
    else:
        parts = ["PM number"]
        if mfr_idx is not None:
            parts.append("manufacturer")
        if model_idx is not None:
            parts.append("model")
        print(f"  Name: auto-composed from {' + '.join(parts)}")
    for label, idx in [
        ("Serial", serial_idx), ("Type", type_idx), ("Slot", slot_idx),
        ("Description", desc_idx), ("Image", image_idx),
        ("Manufacturer", mfr_idx), ("Model", model_idx),
        ("Barcode", barcode_idx), ("Calibration", cal_idx),
    ]:
        if idx is not None:
            print(f"  {label}: '{headers[idx]}' (col {idx + 1})")

    # --- First pass: build locker slot auto-numbering for "schrank*" values ---
    schrank_row_indices = []
    if slot_idx is not None:
        for data_idx, row in enumerate(rows[1:]):
            if row[slot_idx]:
                val = str(row[slot_idx]).strip().lower()
                if val.startswith("schrank"):
                    schrank_row_indices.append(data_idx)
    schrank_slot_map = {idx: slot_num for slot_num, idx in enumerate(schrank_row_indices, start=1)}

    if schrank_slot_map:
        print(f"\n  Auto-numbering {len(schrank_slot_map)} 'schrank' devices → slots 1-{len(schrank_slot_map)}")

    # --- Second pass: parse all data rows ---
    devices = []
    skipped_empty = 0
    for data_idx, row in enumerate(rows[1:]):
        row_num = data_idx + 2  # Excel row number (1-indexed header + 1)

        # PM number (required)
        pm_number = str(row[pm_idx]).strip() if row[pm_idx] else ""
        if not pm_number:
            skipped_empty += 1
            continue

        # Manufacturer
        manufacturer = None
        if mfr_idx is not None and row[mfr_idx]:
            manufacturer = str(row[mfr_idx]).strip()

        # Model
        model_val = None
        if model_idx is not None and row[model_idx]:
            model_val = str(row[model_idx]).strip()

        # Name: compose or read directly
        if compose_name:
            name_parts = [pm_number]
            if manufacturer:
                name_parts.append(manufacturer)
            if model_val:
                name_parts.append(model_val)
            name = " ".join(name_parts)
        else:
            name = str(row[name_idx]).strip() if row[name_idx] else pm_number

        # Serial number (optional)
        serial = None
        if serial_idx is not None and row[serial_idx]:
            serial = str(row[serial_idx]).strip()

        # Device type / category
        device_type = args.default_type
        if type_idx is not None and row[type_idx]:
            device_type = str(row[type_idx]).strip()

        # Locker slot (auto-number schrank or parse integer)
        locker_slot = None
        if slot_idx is not None and row[slot_idx]:
            raw_val = str(row[slot_idx]).strip()
            if raw_val.lower().startswith("schrank"):
                locker_slot = schrank_slot_map.get(data_idx)
            else:
                try:
                    locker_slot = int(raw_val)
                except (ValueError, TypeError):
                    pass

        # Description
        description = None
        if desc_idx is not None and row[desc_idx]:
            description = str(row[desc_idx]).strip()

        # Image path
        image_path = None
        if image_idx is not None and row[image_idx]:
            image_path = str(row[image_idx]).strip()

        # Barcode
        barcode = None
        if barcode_idx is not None and row[barcode_idx]:
            barcode = str(row[barcode_idx]).strip()

        # Calibration due date
        calibration_due = None
        if cal_idx is not None:
            calibration_due = _parse_date(row[cal_idx])

        devices.append({
            "pm_number": pm_number,
            "name": name,
            "serial_number": serial,
            "device_type": device_type,
            "locker_slot": locker_slot,
            "description": description,
            "image_path": image_path,
            "manufacturer": manufacturer,
            "model": model_val,
            "barcode": barcode,
            "calibration_due": calibration_due,
            "row": row_num,
        })

    print(f"\nParsed {len(devices)} devices from {len(rows) - 1} data rows ({skipped_empty} skipped due to empty PM number).")

    if not devices:
        print("Nothing to import.")
        return

    # Preview first 10
    print("\nPreview (first 10):")
    print(f"  {'Row':<5} {'PM':<15} {'Name':<35} {'Type':<18} {'Slot':<5} {'Cal. Due'}")
    print(f"  {'---':<5} {'---':<15} {'---':<35} {'---':<18} {'---':<5} {'---'}")
    for d in devices[:10]:
        slot_str = str(d["locker_slot"]) if d["locker_slot"] is not None else "-"
        cal_str = d["calibration_due"].isoformat() if d["calibration_due"] else "-"
        print(f"  {d['row']:<5} {d['pm_number']:<15} {d['name'][:34]:<35} {d['device_type'][:17]:<18} {slot_str:<5} {cal_str}")
    if len(devices) > 10:
        print(f"  ... and {len(devices) - 10} more")

    if args.dry_run:
        print("\n[DRY RUN] No changes written to database.")
        return

    # Import to database
    init_db()
    imported = 0
    duplicates = 0
    errors = 0

    with get_session() as session:
        # Get existing PM numbers to skip duplicates
        existing = DeviceRepository.list_all(session)
        existing_pms = {d.pm_number for d in existing if d.pm_number}

        for d in devices:
            if d["pm_number"] in existing_pms:
                duplicates += 1
                continue
            try:
                DeviceRepository.create(
                    session,
                    name=d["name"],
                    device_type=d["device_type"],
                    pm_number=d["pm_number"],
                    serial_number=d["serial_number"],
                    locker_slot=d["locker_slot"],
                    description=d["description"],
                    image_path=d["image_path"],
                    manufacturer=d["manufacturer"],
                    model=d["model"],
                    barcode=d["barcode"],
                    calibration_due=d["calibration_due"],
                )
                existing_pms.add(d["pm_number"])
                imported += 1
            except Exception as e:
                errors += 1
                print(f"  ERROR row {d['row']}: {e}")

    print(f"\nDone: {imported} imported, {duplicates} duplicates skipped, {errors} errors.")

    # Sync to Excel
    if imported > 0:
        from config.settings import EXCEL_SYNC_PATH
        from smart_locker.database.engine import get_engine
        from smart_locker.sync.excel_sync import export_to_excel
        export_to_excel(get_engine(), EXCEL_SYNC_PATH)
        print(f"Excel sync: {EXCEL_SYNC_PATH}")


if __name__ == "__main__":
    main()
