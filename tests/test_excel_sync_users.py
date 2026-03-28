"""Tests for the Users sheet in Excel auto-sync."""

import tempfile
from pathlib import Path

import pytest
from openpyxl import load_workbook

from smart_locker.database.engine import get_engine
from smart_locker.database.repositories import UserRepository
from smart_locker.sync.excel_sync import export_to_excel


class TestUsersSheet:
    def test_export_includes_users_sheet(self, db_session):
        """Excel export contains a Users sheet with registered users."""
        UserRepository.create(
            db_session,
            display_name="Alice",
            uid_hmac="a" * 64,
            encrypted_card_uid="encrypted_alice",
            role="admin",
        )
        UserRepository.create(
            db_session,
            display_name="Bob",
            uid_hmac="b" * 64,
            encrypted_card_uid="encrypted_bob",
            role="user",
        )
        db_session.commit()

        path = Path(tempfile.mktemp(suffix=".xlsx"))
        try:
            export_to_excel(get_engine(), path)
            wb = load_workbook(path, read_only=True)

            assert "Users" in wb.sheetnames
            ws = wb["Users"]
            rows = list(ws.iter_rows(values_only=True))

            # Header row
            assert rows[0] == ("ID", "Display Name", "Role", "Active", "Registered At")

            # Data rows (sorted by display_name)
            assert len(rows) == 3  # header + 2 users
            assert rows[1][1] == "Alice"
            assert rows[1][2] == "admin"
            assert rows[1][3] == "Yes"
            assert rows[2][1] == "Bob"
            assert rows[2][2] == "user"
            wb.close()
        finally:
            path.unlink(missing_ok=True)

    def test_no_sensitive_data_in_users_sheet(self, db_session):
        """Users sheet must not contain HMAC, encrypted UID, or card data."""
        UserRepository.create(
            db_session,
            display_name="Charlie",
            uid_hmac="c" * 64,
            encrypted_card_uid="encrypted_charlie",
        )
        db_session.commit()

        path = Path(tempfile.mktemp(suffix=".xlsx"))
        try:
            export_to_excel(get_engine(), path)
            wb = load_workbook(path, read_only=True)
            ws = wb["Users"]
            rows = list(ws.iter_rows(values_only=True))

            headers_lower = [str(h).lower() for h in rows[0]]
            for forbidden in ["hmac", "encrypted", "uid", "card"]:
                assert not any(forbidden in h for h in headers_lower), \
                    f"Header should not contain '{forbidden}'"

            # Check data cells don't contain the actual HMAC or encrypted values
            for row in rows[1:]:
                for cell in row:
                    cell_str = str(cell) if cell else ""
                    assert "c" * 64 not in cell_str
                    assert "encrypted_charlie" not in cell_str
            wb.close()
        finally:
            path.unlink(missing_ok=True)

    def test_empty_users(self, db_session):
        """Users sheet still has a header row when no users exist."""
        path = Path(tempfile.mktemp(suffix=".xlsx"))
        try:
            export_to_excel(get_engine(), path)
            wb = load_workbook(path, read_only=True)
            ws = wb["Users"]
            rows = list(ws.iter_rows(values_only=True))
            assert len(rows) == 1  # header only
            assert rows[0][1] == "Display Name"
            wb.close()
        finally:
            path.unlink(missing_ok=True)
