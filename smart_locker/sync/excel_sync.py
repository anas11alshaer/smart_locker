"""
File: excel_sync.py
Description: On-demand Excel export of device, transaction, and user data.
             Provides two export modes: ``export_to_excel()`` writes a styled
             three-sheet workbook to disk (used by CLI scripts like
             ``update_device.py``), and ``export_to_excel_bytes()`` returns
             the same workbook as raw bytes (used by the admin download
             endpoint to serve the file as an HTTP response without touching
             disk).
Project: smart_locker/sync
Notes: Both public functions delegate to ``_build_workbook()`` which queries
       the database and assembles the openpyxl Workbook in memory. The disk-
       writing path uses a temp file + atomic replace for crash safety, with
       retry logic to handle Windows file locking when the target is open in
       Excel. Three sheets are produced: Devices, Transactions, and Users.
"""

import io
import logging
import os
import tempfile
import time
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from sqlalchemy import select
from sqlalchemy.orm import Session

from smart_locker.database.models import Device, TransactionLog, User

logger = logging.getLogger(__name__)


def _build_workbook(engine) -> Workbook:
    """Query the database and build a styled three-sheet Excel workbook.

    Creates sheets for Devices (inventory with status and borrower),
    Transactions (borrow/return history in reverse chronological order),
    and Users (registered users with role and registration date). Each
    sheet has bold blue headers and auto-sized column widths.

    This is a private helper — callers should use ``export_to_excel()``
    (disk) or ``export_to_excel_bytes()`` (in-memory bytes) instead.

    Args:
        engine: SQLAlchemy Engine instance to query data from.

    Returns:
        An openpyxl Workbook ready to be saved to disk or serialised
        to bytes.
    """
    with Session(engine) as session:
        devices = session.execute(
            select(Device).order_by(Device.locker_slot, Device.name)
        ).scalars().all()

        transactions = session.execute(
            select(TransactionLog).order_by(TransactionLog.timestamp.desc())
        ).scalars().all()

        users = session.execute(
            select(User).order_by(User.display_name)
        ).scalars().all()

        wb = Workbook()

        # Shared header styling — bold white text on steel-blue background
        header_font = Font(bold=True)
        header_fill = PatternFill(
            start_color="D9E1F2", end_color="D9E1F2", fill_type="solid"
        )

        # --- Devices sheet ---------------------------------------------------
        ws = wb.active
        ws.title = "Devices"
        headers = [
            "PM Number", "Name", "Type", "Manufacturer", "Model",
            "Serial Number", "Barcode", "Locker Slot", "Status",
            "Current Borrower", "Description", "Calibration Due",
        ]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = header_fill

        for d in devices:
            borrower = ""
            if d.current_borrower is not None:
                borrower = d.current_borrower.display_name
            ws.append([
                d.pm_number,
                d.name,
                d.device_type,
                d.manufacturer,
                d.model,
                d.serial_number,
                d.barcode,
                d.locker_slot,
                d.status.value,
                borrower,
                d.description,
                d.calibration_due,
            ])

        _auto_width(ws)

        # --- Transactions sheet -----------------------------------------------
        ws2 = wb.create_sheet("Transactions")
        txn_headers = [
            "Date", "User", "Device", "Type", "Performed By", "Notes",
        ]
        ws2.append(txn_headers)
        for cell in ws2[1]:
            cell.font = header_font
            cell.fill = header_fill

        for t in transactions:
            user_name = t.user.display_name if t.user else ""
            device_name = t.device.name if t.device else ""
            admin_name = t.performed_by.display_name if t.performed_by else ""
            ws2.append([
                t.timestamp.strftime("%Y-%m-%d %H:%M:%S") if t.timestamp else "",
                user_name,
                device_name,
                t.transaction_type.value,
                admin_name,
                t.notes,
            ])

        _auto_width(ws2)

        # --- Users sheet ------------------------------------------------------
        ws3 = wb.create_sheet("Users")
        user_headers = [
            "ID", "Display Name", "Role", "Active", "Registered At",
        ]
        ws3.append(user_headers)
        for cell in ws3[1]:
            cell.font = header_font
            cell.fill = header_fill

        for u in users:
            ws3.append([
                u.id,
                u.display_name,
                u.role.value,
                "Yes" if u.is_active else "No",
                u.created_at.strftime("%Y-%m-%d %H:%M:%S") if u.created_at else "",
            ])

        _auto_width(ws3)

    return wb


def _auto_width(ws) -> None:
    """Auto-size every column in a worksheet based on cell content length.

    Iterates all columns, measures the longest string value, and sets
    the column width to that length plus padding (capped at 40
    characters to prevent excessively wide columns).

    Args:
        ws: An openpyxl Worksheet to resize.

    Returns:
        None. Column widths are modified in place.
    """
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            val = str(cell.value) if cell.value is not None else ""
            max_len = max(max_len, len(val))
        # +3 for padding, capped at 40 to prevent overly wide columns
        ws.column_dimensions[col_letter].width = min(max_len + 3, 40)


def export_to_excel(engine, output_path: str | Path) -> None:
    """Export current data to an Excel file on disk.

    Builds the workbook via ``_build_workbook()``, writes it to a
    temporary file in the same directory as the target, then atomically
    replaces the target. If the target is locked (e.g., open in Excel),
    retries up to 3 times with a 1-second delay before giving up.

    Used by CLI scripts (``scripts/update_device.py``) that need to
    produce a file on disk after modifying the database.

    Args:
        engine: SQLAlchemy Engine instance to query data from.
        output_path: Destination file path for the Excel workbook.

    Returns:
        None. The Excel file is written to disk at ``output_path``.
    """
    path = Path(output_path)
    wb = _build_workbook(engine)

    # Write to a temp file first, then atomically replace the target.
    # This avoids partial writes if the process is interrupted mid-save.
    tmp_fd, tmp_path_str = tempfile.mkstemp(suffix=".xlsx", dir=path.parent)
    os.close(tmp_fd)
    tmp_path = Path(tmp_path_str)

    try:
        wb.save(tmp_path)

        # Retry the atomic replace up to 3 times — the file may be
        # momentarily locked by Excel (e.g., during an auto-save cycle)
        # but released within a second or two.
        max_retries = 3
        retry_delay_seconds = 1.0
        for attempt in range(1, max_retries + 1):
            try:
                tmp_path.replace(path)
                logger.debug("Excel export: %s written successfully", path)
                # Replace succeeded — exit the retry loop
                break
            except PermissionError:
                if attempt < max_retries:
                    # File is locked — wait briefly and retry in case
                    # Excel releases the lock (e.g., after auto-save)
                    logger.debug(
                        "Excel export: %s is locked, retrying in %ss "
                        "(attempt %d/%d)",
                        path, retry_delay_seconds, attempt, max_retries,
                    )
                    time.sleep(retry_delay_seconds)
                else:
                    # All retries exhausted — log warning and give up
                    logger.warning(
                        "Excel export: %s is locked (open in Excel). "
                        "Export skipped.",
                        path,
                    )
    finally:
        # Always clean up the temp file — never leave orphaned files in
        # the output directory regardless of success or failure
        if tmp_path.exists():
            try:
                tmp_path.unlink()
            except OSError:
                # Temp file removal is best-effort; may fail if another
                # process grabbed it, but this is not a critical error
                pass


def export_to_excel_bytes(engine) -> bytes:
    """Export current data as an in-memory Excel file (raw bytes).

    Builds the same three-sheet workbook as ``export_to_excel()`` but
    serialises it to a ``BytesIO`` buffer instead of writing to disk.
    This avoids all file-locking issues and is used by the admin
    download endpoint to return the workbook as an HTTP response.

    Args:
        engine: SQLAlchemy Engine instance to query data from.

    Returns:
        Raw bytes of the ``.xlsx`` file, suitable for streaming in an
        HTTP response with content-type
        ``application/vnd.openxmlformats-officedocument.spreadsheetml.sheet``.
    """
    wb = _build_workbook(engine)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
