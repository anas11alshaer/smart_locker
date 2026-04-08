"""
File: source_import.py
Description: Source Excel import — reads the company device master list and syncs
             to the database. Inserts new devices and updates metadata on existing
             ones. Only devices with a slot column starting with "schrank" are
             imported. Supports both German and English column headers.
Project: smart_locker/sync
Notes: Can be called programmatically (by the scheduler), via CLI
       ('python -m scripts.sync_source'), or via the admin API endpoint
       (POST /api/admin/sync-source). Status, borrower, slot, image, and
       description fields are never overwritten on existing devices.
"""

import logging
import shutil
import tempfile
from dataclasses import dataclass, field
from datetime import date, datetime
from pathlib import Path

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from smart_locker.database.models import Device, DeviceStatus
from smart_locker.database.repositories import DeviceRepository

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Column auto-detection candidates (German + English)
# ---------------------------------------------------------------------------

PM_CANDIDATES = [
    "equipment", "pm number", "pm", "pm_number", "equipment number",
    "equipmentnumber", "inventarnummer",
]
NAME_CANDIDATES = [
    "name", "device name", "device", "equipment name", "item", "item name",
]
SERIAL_CANDIDATES = [
    "serial", "serial number", "s/n", "sn", "serial no", "serial_number",
    "serialnumber", "hersteller-serialnummer", "hersteller-seriennummer",
    "herstellerseriennummer", "seriennummer",
]
TYPE_CANDIDATES = [
    "type", "device type", "category", "device_type", "kind", "kategorie",
]
SLOT_CANDIDATES = [
    "slot", "locker slot", "locker", "locker_slot", "bay",
    "platz messmittelschrank", "platz", "messmittelschrank", "schrank",
]
DESC_CANDIDATES = [
    "description", "desc", "details", "notes", "beschreibung",
]
IMAGE_CANDIDATES = [
    "image", "photo", "image_path", "photo_path", "img", "picture", "filename",
]
MANUFACTURER_CANDIDATES = [
    "manufacturer", "make", "brand", "hersteller",
]
MODEL_CANDIDATES = [
    "model", "model name", "type designation",
    "typbezeichnung", "typ", "modell",
]
BARCODE_CANDIDATES = [
    "barcode", "bar code", "barcodenummer",
]
CALIBRATION_CANDIDATES = [
    "calibration due", "calibration_due", "next calibration",
    "datum der nächsten kalibrierung", "datum der nachsten kalibrierung",
    "datum der n\u00e4chsten kalibrierung",
    "nächste kalibrierung", "kalibrierung", "kalibrierdatum",
]
LOCATION_CANDIDATES = [
    "aktueller einsatzort",
]


# ---------------------------------------------------------------------------
# Result dataclass
# ---------------------------------------------------------------------------

@dataclass
class ImportResult:
    """Summary of a source import run with per-category counts.

    Tracks how many devices were newly imported, updated with changed
    metadata, left unchanged, skipped (non-locker), or errored during
    the import process. Also records how many new registrant names were
    extracted from the "Aktueller Einsatzort" column and added to the
    registrants table for self-service registration.
    """
    imported: int = 0
    updated: int = 0
    unchanged: int = 0
    non_locker_skipped: int = 0
    errors: int = 0
    error_details: list[str] = field(default_factory=list)
    registrants_added: int = 0


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def find_column(headers: list[str], candidates: list[str]) -> int | None:
    """Find a column index by trying multiple possible header names (case-insensitive).

    Iterates over the header row and returns the index of the first header
    whose stripped, lowercased text matches any of the candidate names.

    Args:
        headers: List of column header strings from the Excel file.
        candidates: Possible header names to match against (e.g. German
                    and English variants).

    Returns:
        Zero-based column index if found, or None if no match.
    """
    lower_candidates = [c.lower() for c in candidates]
    for i, header in enumerate(headers):
        if header and header.strip().lower() in lower_candidates:
            return i
    return None


def parse_date(value) -> date | None:
    """Parse a date from an Excel cell value (datetime object or string).

    Handles native ``datetime``/``date`` objects (common in openpyxl) and
    string values in DD.MM.YYYY, YYYY-MM-DD, or DD/MM/YYYY format.

    Args:
        value: Cell value from openpyxl — may be datetime, date, str, or None.

    Returns:
        A ``date`` object, or None if the value is empty or unparseable.
    """
    if value is None:
        return None
    if isinstance(value, date) and not isinstance(value, datetime):
        return value
    if isinstance(value, datetime):
        return value.date()
    text = str(value).strip()
    if not text:
        return None
    for fmt in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _detect_columns(
    headers: list[str],
    overrides: dict[str, str] | None = None,
) -> dict[str, int | None]:
    """Detect column indices from headers, with optional manual overrides.

    For each field (pm, name, serial, type, slot, etc.), tries to match the
    override value first (if provided), then falls back to the predefined
    candidate lists for auto-detection.

    Args:
        headers: List of column header strings from the Excel file.
        overrides: Optional mapping of field names to explicit header strings
                   (e.g. ``{"pm": "Equipment Nr"}``).

    Returns:
        Dict mapping field names to their detected column index (or None).
    """
    ov = overrides or {}
    return {
        "pm":           find_column(headers, [ov["pm"]] if ov.get("pm") else PM_CANDIDATES),
        "name":         find_column(headers, [ov["name"]] if ov.get("name") else NAME_CANDIDATES),
        "serial":       find_column(headers, [ov["serial"]] if ov.get("serial") else SERIAL_CANDIDATES),
        "type":         find_column(headers, [ov["type"]] if ov.get("type") else TYPE_CANDIDATES),
        "slot":         find_column(headers, [ov["slot"]] if ov.get("slot") else SLOT_CANDIDATES),
        "desc":         find_column(headers, DESC_CANDIDATES),
        "image":        find_column(headers, IMAGE_CANDIDATES),
        "manufacturer": find_column(headers, [ov["manufacturer"]] if ov.get("manufacturer") else MANUFACTURER_CANDIDATES),
        "model":        find_column(headers, [ov["model"]] if ov.get("model") else MODEL_CANDIDATES),
        "barcode":      find_column(headers, [ov["barcode"]] if ov.get("barcode") else BARCODE_CANDIDATES),
        "calibration":  find_column(headers, [ov["calibration"]] if ov.get("calibration") else CALIBRATION_CANDIDATES),
        "location":     find_column(headers, [ov["location"]] if ov.get("location") else LOCATION_CANDIDATES),
    }


def _cell_str(row, idx: int | None) -> str | None:
    """Read a cell as a stripped string, or None.

    Args:
        row: A tuple of cell values from openpyxl (one row of data).
        idx: Column index to read, or None to skip.

    Returns:
        The cell value as a stripped string, or None if the index is None,
        the cell is None, or the stripped string is empty.
    """
    if idx is None or row[idx] is None:
        return None
    val = str(row[idx]).strip()
    return val if val else None


# ---------------------------------------------------------------------------
# Main import function
# ---------------------------------------------------------------------------

def import_from_source_excel(
    engine,
    source_path: str | Path,
    sheet_name: str | None = None,
    dry_run: bool = False,
    default_type: str = "general",
    column_overrides: dict[str, str] | None = None,
) -> ImportResult:
    """Import new devices and update existing ones from the company source Excel.

    Only devices with a slot column value starting with "schrank" are imported.
    Existing devices (matched by PM number) get metadata updated; status, borrower,
    locker_slot, image_path, and description are never overwritten.

    Args:
        engine: SQLAlchemy engine.
        source_path: Path to the .xlsx file.
        sheet_name: Specific sheet to read (default: active sheet).
        dry_run: If True, parse and report but don't write to DB.
        default_type: Default device_type when no type column is found.
        column_overrides: Dict mapping field names to header strings for manual column mapping.

    Returns:
        ImportResult with counts.
    """
    result = ImportResult()
    path = Path(source_path)

    if not path.exists():
        logger.warning("Source Excel not found: %s", path)
        result.errors = 1
        result.error_details.append(f"File not found: {path}")
        return result

    logger.info("Reading source Excel: %s", path)

    # Copy to a temp file before reading so the import succeeds even when
    # the source file is open in Excel (Windows holds a lock on open files).
    tmp_fd, tmp_path_str = tempfile.mkstemp(suffix=".xlsx")
    tmp_path = Path(tmp_path_str)
    try:
        shutil.copy2(path, tmp_path)
    except PermissionError:
        logger.warning(
            "Source Excel at %s is locked — cannot copy for reading.", path
        )
        result.errors = 1
        result.error_details.append(f"Source file locked: {path}")
        tmp_path.unlink(missing_ok=True)
        return result
    finally:
        # Close the file descriptor opened by mkstemp; the file itself
        # persists on disk for openpyxl to read.
        import os
        os.close(tmp_fd)

    try:
        wb = load_workbook(tmp_path, read_only=True, data_only=True)

        if sheet_name:
            if sheet_name not in wb.sheetnames:
                result.errors = 1
                result.error_details.append(
                    f"Sheet '{sheet_name}' not found. Available: {wb.sheetnames}"
                )
                wb.close()
                return result
            ws = wb[sheet_name]
        else:
            ws = wb.active

        # Read all rows
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
    finally:
        # Always clean up the temporary copy
        tmp_path.unlink(missing_ok=True)

    if len(rows) < 2:
        logger.warning("Source Excel has no data rows.")
        return result

    # Parse headers
    headers = [str(h).strip() if h else "" for h in rows[0]]
    cols = _detect_columns(headers, column_overrides)

    if cols["pm"] is None:
        result.errors = 1
        result.error_details.append(f"Could not find PM/equipment column. Headers: {headers}")
        return result

    slot_idx = cols["slot"]
    pm_idx = cols["pm"]
    compose_name = cols["name"] is None

    # --- Registrant extraction: collect unique person names from ALL rows ---
    # The "Aktueller Einsatzort" column lists where each device is deployed.
    # Values containing "schrank" are locker locations; everything else is a
    # person's name. We read ALL rows (not just schrank-filtered ones) so that
    # the registrant list covers employees who have non-locker devices too.
    registrant_names: set[str] = set()
    if cols["location"] is not None:
        for row in rows[1:]:
            location = _cell_str(row, cols["location"])
            if location and "schrank" not in location.lower():
                registrant_names.add(location.strip())

    if registrant_names:
        logger.info(
            "Found %d unique registrant name(s) in 'Aktueller Einsatzort' column.",
            len(registrant_names),
        )

    # --- First pass: build schrank auto-numbering ---
    schrank_row_indices: list[int] = []
    if slot_idx is not None:
        for data_idx, row in enumerate(rows[1:]):
            if row[slot_idx]:
                val = str(row[slot_idx]).strip().lower()
                if val.startswith("schrank"):
                    schrank_row_indices.append(data_idx)

    schrank_slot_map = {idx: slot_num for slot_num, idx in enumerate(schrank_row_indices, start=1)}

    # --- Second pass: parse rows, filter to schrank only ---
    parsed_devices: list[dict] = []
    for data_idx, row in enumerate(rows[1:]):
        pm_number = _cell_str(row, pm_idx)
        if not pm_number:
            continue

        # Only import devices assigned to a locker
        if slot_idx is None:
            result.non_locker_skipped += 1
            continue
        raw_slot = str(row[slot_idx]).strip().lower() if row[slot_idx] else ""
        if not raw_slot.startswith("schrank"):
            result.non_locker_skipped += 1
            continue

        manufacturer = _cell_str(row, cols["manufacturer"])
        model_val = _cell_str(row, cols["model"])

        if compose_name:
            # Compose name from manufacturer and model only — PM number is a
            # separate identifier and should not appear in the display name.
            name_parts = []
            if manufacturer:
                name_parts.append(manufacturer)
            if model_val:
                name_parts.append(model_val)
            # Fall back to PM number only if neither manufacturer nor model exists
            name = " ".join(name_parts) if name_parts else pm_number
        else:
            name = _cell_str(row, cols["name"]) or pm_number

        serial = _cell_str(row, cols["serial"])

        device_type = default_type
        if cols["type"] is not None and row[cols["type"]]:
            device_type = str(row[cols["type"]]).strip()

        locker_slot = schrank_slot_map.get(data_idx)

        description = _cell_str(row, cols["desc"])
        image_path = _cell_str(row, cols["image"])
        barcode = _cell_str(row, cols["barcode"])

        calibration_due = None
        if cols["calibration"] is not None:
            calibration_due = parse_date(row[cols["calibration"]])

        # Determine device location from "Aktueller Einsatzort" column.
        # If the value contains "schrank" the device is physically in the
        # locker (AVAILABLE); any other non-empty value is a person's name,
        # meaning that person currently has the device (BORROWED).
        location_raw = _cell_str(row, cols["location"])
        if location_raw and "schrank" not in location_raw.lower():
            borrower_name = location_raw
            device_status = DeviceStatus.BORROWED.value
        else:
            borrower_name = None
            device_status = DeviceStatus.AVAILABLE.value

        parsed_devices.append({
            "pm_number": pm_number,
            "name": name,
            "serial_number": serial,
            "device_type": device_type,
            "locker_slot": locker_slot,
            "description": description,
            "image_path": image_path,
            "manufacturer": manufacturer,
            "model": model_val,
            "barcode": barcode,
            "calibration_due": calibration_due,
            "borrower_name": borrower_name,
            "device_status": device_status,
        })

    logger.info(
        "Parsed %d locker devices (%d non-locker skipped).",
        len(parsed_devices), result.non_locker_skipped,
    )

    if not parsed_devices or dry_run:
        return result

    # --- Import to database ---
    from smart_locker.database.engine import get_session
    from smart_locker.database.repositories import UserRepository
    from smart_locker.sync.excel_sync import export_to_excel

    with get_session() as session:
        for d in parsed_devices:
            try:
                # Resolve the borrower name (from "Aktueller Einsatzort")
                # to a user ID.  If the name doesn't match any registered
                # user, the device is still marked BORROWED but without a
                # linked borrower so the kiosk UI reflects that it is not
                # physically in the locker.
                borrower_id = None
                if d["borrower_name"]:
                    user = UserRepository.find_by_display_name(
                        session, d["borrower_name"]
                    )
                    if user is not None:
                        borrower_id = user.id
                    else:
                        logger.warning(
                            "Borrower '%s' for PM %s not found in user database.",
                            d["borrower_name"], d["pm_number"],
                        )

                existing = DeviceRepository.find_by_pm(session, d["pm_number"])
                if existing is None:
                    # New device — insert
                    DeviceRepository.create(
                        session,
                        name=d["name"],
                        device_type=d["device_type"],
                        pm_number=d["pm_number"],
                        serial_number=d["serial_number"],
                        locker_slot=d["locker_slot"],
                        description=d["description"],
                        image_path=d["image_path"],
                        manufacturer=d["manufacturer"],
                        model=d["model"],
                        barcode=d["barcode"],
                        calibration_due=d["calibration_due"],
                        status=d["device_status"],
                        current_borrower_id=borrower_id,
                    )
                    result.imported += 1
                else:
                    # Existing device — update metadata and location status
                    changed = DeviceRepository.update_metadata(
                        session,
                        existing,
                        name=d["name"],
                        device_type=d["device_type"],
                        serial_number=d["serial_number"],
                        manufacturer=d["manufacturer"],
                        model=d["model"],
                        barcode=d["barcode"],
                        calibration_due=d["calibration_due"],
                        status=DeviceStatus(d["device_status"]),
                        current_borrower_id=borrower_id,
                    )
                    if changed:
                        result.updated += 1
                    else:
                        result.unchanged += 1
            except Exception as e:
                result.errors += 1
                result.error_details.append(f"PM {d['pm_number']}: {e}")
                logger.error("Import error for PM %s: %s", d["pm_number"], e)

    # --- Sync registrant names to the registrants table ---
    # This is done after the device import so both operations share the
    # same get_session factory. Names are additive — existing registrants
    # are never removed, only new ones are inserted.
    if registrant_names:
        from smart_locker.database.repositories import RegistrantRepository

        try:
            with get_session() as reg_session:
                added = RegistrantRepository.add_names(reg_session, registrant_names)
                result.registrants_added = added
        except Exception as e:
            logger.warning("Registrant name sync failed: %s", e)

    # Trigger output Excel sync
    if result.imported > 0 or result.updated > 0:
        try:
            export_to_excel(engine)
        except Exception as e:
            logger.warning("Excel sync after import failed: %s", e)

    logger.info(
        "Source import done: %d imported, %d updated, %d unchanged, %d errors, "
        "%d registrants added.",
        result.imported, result.updated, result.unchanged, result.errors,
        result.registrants_added,
    )
    return result
